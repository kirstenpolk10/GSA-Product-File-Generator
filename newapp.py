import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from io import BytesIO, StringIO
import os
import requests
from collections import Counter
import tempfile
import shutil
import zipfile
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- Config ---
ALLOWED_SINS = {
    "332510C",
    "332999",
    "333316P",
    "33411",
    "339920S",
    "511130",
    "OLM"
}

FILE_CONFIG = {
    "start_row": 3,
    "row_step": 2,
    "column_map": {
        "Part Number": 2,
        "Item Name": 3,
        "Manufacturer": 4,
        "Unit Price": 6,
        "Extended Price": 8,
        "Sin Number": 10,
        "Description": 12
    },
    "template_map": {
        "Manufacturer": 2,
        "Part Number": (3, 4),
        "Sin Number": 5,  # Column E
        "Item Name": 6,
        "Description": 7,
        "Total Sales": 32,
        "Unit Price": 16
    },
    "max_rows_per_file": 10000,
    "output_base_filename": "FCP_Product_File",
    "default_font": Font(name='Calibri', size=12)
}

TEMPLATE_URL = "https://raw.githubusercontent.com/kirstenpolk10/GSA-Product-File-Generator/master/FCP_Product_FileBlankTemplate.xlsx"
SIN_MAPPING_URL = "https://raw.githubusercontent.com/kirstenpolk10/GSA-Product-File-Generator/master/sin_group_mapping.csv"

@st.cache_data
def load_template_from_github():
    response = requests.get(TEMPLATE_URL)
    response.raise_for_status()
    return BytesIO(response.content)

@st.cache_data
def load_sin_mapping_from_github():
    response = requests.get(SIN_MAPPING_URL)
    response.raise_for_status()
    df = pd.read_csv(BytesIO(response.content), header=None)
    return dict(zip(df.iloc[:, 0].astype(str), df.iloc[:, 1].astype(str)))

def clean_description(description):
    if not isinstance(description, str):
        description = str(description)
    return description.strip()[:40]

def extract_data(df):
    data = []
    start_idx = FILE_CONFIG["start_row"] - 1
    for i in range(start_idx, len(df), FILE_CONFIG["row_step"]):
        row = df.iloc[i]
        record = {
            field: row[idx] if idx < len(row) else ""
            for field, idx in FILE_CONFIG["column_map"].items()
        }
        record["Description"] = clean_description(record["Description"])
        data.append(record)
    return data

def write_to_template(ws, row_num, record):
    font = FILE_CONFIG["default_font"]
    ws.cell(row=row_num, column=1, value="B").font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Manufacturer"], value=record["Manufacturer"]).font = font
    part_num = record["Part Number"]
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Part Number"][0], value=part_num).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Part Number"][1], value=part_num).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Item Name"], value=record["Item Name"]).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Description"], value=record["Description"]).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Unit Price"], value=record["Unit Price"]).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Total Sales"], value=record["Extended Price"]).font = font
    ws.cell(row=row_num, column=FILE_CONFIG["template_map"]["Sin Number"], value=record["Sin Number"]).font = font

    ws.cell(row=row_num, column=12, value=f"=P{row_num}*1.4").font = font
    ws.cell(row=row_num, column=15, value=f"=P{row_num}*0.9925").font = font

    for col, val in zip([17, 18, 19, 20, 21, 22, 23, 30], ["MX", "15", "AE", "D", "O", "O", "O", "AlegnaLogo.jpg"]):
        ws.cell(row=row_num, column=col, value=val).font = font

def process_all(uploaded_files):
    with st.spinner("Processing files..."):
        temp_dir = tempfile.mkdtemp()
        output_dir = os.path.join(temp_dir, "output")
        os.makedirs(output_dir, exist_ok=True)

        template_file = load_template_from_github()
        sin_mapping = load_sin_mapping_from_github()

        wb = ws = None
        current_row = 3
        file_counter = 1
        report = []
        current_output_path = None
        batch_first_file = uploaded_files[0].name
        batch_last_file = uploaded_files[0].name

        def create_new_output_file():
            nonlocal wb, ws, current_row, current_output_path
            output_path = os.path.join(output_dir, f"{FILE_CONFIG['output_base_filename']}_{file_counter}_temp.xlsx")
            with open(output_path, "wb") as f:
                f.write(template_file.getbuffer())
            wb = load_workbook(output_path)
            ws = wb.active
            current_row = 3
            current_output_path = output_path
            return output_path

        def finalize_output_file(output_path, first_file, last_file):
            nonlocal wb
            wb.save(output_path)
            def code_from_filename(filename):
                base = os.path.splitext(filename)[0]
                return base[:2].upper() if len(base) >= 2 else "XX"
            first_code = code_from_filename(first_file)
            last_code = code_from_filename(last_file)
            final_name = f"{FILE_CONFIG['output_base_filename']}_{file_counter}({first_code}-{last_code}).xlsx"
            final_path = os.path.join(output_dir, final_name)
            os.rename(output_path, final_path)
            return final_path

        current_output_path = create_new_output_file()

        for uploaded in uploaded_files:
            file = uploaded.name
            st.write(f"🔍 Processing: {file}")
            try:
                if file.lower().endswith('.csv'):
                    df = pd.read_csv(uploaded, header=None, low_memory=False)
                else:
                    df = pd.read_excel(uploaded, header=None)

                records = extract_data(df)

                # --- Count SINs ---
                sin_counts = Counter(str(r.get("Sin Number", "")).strip() for r in records if r.get("Sin Number"))
                most_common = [sin for sin, _ in sin_counts.most_common()]

                final_sin = ""
                for sin in most_common:
                    if sin in ALLOWED_SINS:
                        final_sin = sin
                        break
                if not final_sin:
                    for sin in most_common:
                        mapped = sin_mapping.get(sin)
                        if mapped in ALLOWED_SINS:
                            final_sin = mapped
                            break

                if not final_sin:
                    report.append(f"⚠️ No valid SIN found for {file}. Skipping.")
                    continue

                for r in records:
                    r["Sin Number"] = final_sin

                for record in records:
                    batch_last_file = file
                    if current_row > FILE_CONFIG["max_rows_per_file"] + 2:
                        finalize_output_file(current_output_path, batch_first_file, batch_last_file)
                        file_counter += 1
                        batch_first_file = file
                        current_output_path = create_new_output_file()

                    write_to_template(ws, current_row, record)
                    current_row += 1

                report.append(f"✅ Successfully processed {file} using SIN: {final_sin}")

            except Exception as e:
                report.append(f"❌ Failed to process {file}: {str(e)}")

        if current_output_path:
            finalize_output_file(current_output_path, batch_first_file, batch_last_file)

        report_path = os.path.join(output_dir, "Processing_Report.txt")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(report))

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for file in os.listdir(output_dir):
                zipf.write(os.path.join(output_dir, file), file)
        zip_buffer.seek(0)

        st.success("✅ Processing complete!")
        st.download_button(
            "📥 Download Results (ZIP)",
            zip_buffer,
            file_name="Processed_Results.zip",
            mime="application/zip"
        )

        with st.expander("Processing Report"):
            st.text("\n".join(report))

        shutil.rmtree(temp_dir)

# --- Streamlit UI ---
st.set_page_config(page_title="FCP Generator", layout="centered")
st.title("📦 FCP Product File Generator")

uploaded_files = st.file_uploader("Upload input files (Excel or CSV)", type=["xlsx", "xls", "csv"], accept_multiple_files=True)
if uploaded_files:
    if st.button("Process Files"):
        process_all(uploaded_files)
else:
    st.info("Please upload at least one Excel or CSV file to begin.")