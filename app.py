import os
import re
import pandas as pd
import shutil
import tempfile
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font
import streamlit as st
import zipfile

warnings = __import__("warnings")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------- Streamlit UI ----------------------
st.title("📦 FCP Product File Generator")

uploaded_files = st.file_uploader("Upload input files (Excel or CSV)", type=["xlsx", "xls", "csv"], accept_multiple_files=True)

template_file = st.file_uploader("Upload template file (Excel)", type=["xlsx"])
sin_mapping_file = st.file_uploader("Upload SIN mapping file", type=["xlsx"])

process_button = st.button("Process Files")

# ---------------------- Process Logic ----------------------
if process_button and uploaded_files and template_file and sin_mapping_file:
    with st.spinner("Processing files..."):
        # Temp directories
        temp_dir = tempfile.mkdtemp()
        output_dir = os.path.join(temp_dir, "output")
        os.makedirs(output_dir, exist_ok=True)

        # Constants
        max_rows_per_file = 10000
        output_base_filename = "FCP_Product_File"
        column_names = [
            "A", "Manufacturer", "Part Number", "Description", "Total Sales",
            "E", "F", "G", "H", "Max List Price", "K", "L", "M", "N", "O"
        ]
        default_font = Font(name='Calibri', size=12)

        # SIN Mapping
        sin_df = pd.read_excel(sin_mapping_file, header=None)
        sin_dict = dict(zip(sin_df.iloc[:, 0].astype(str), sin_df.iloc[:, 1]))

        # Template save
        template_path = os.path.join(temp_dir, "template.xlsx")
        with open(template_path, "wb") as f:
            f.write(template_file.read())

        # Output logic
        report = []
        file_counter = 1
        current_row = 3
        combined_output_path = ""
        wb = None
        ws = None

        def create_new_output_file():
            global wb, ws, current_row, combined_output_path
            temp_name = f"{output_base_filename}_{file_counter}_temp.xlsx"
            combined_output_path = os.path.join(output_dir, temp_name)
            shutil.copy(template_path, combined_output_path)
            wb = load_workbook(combined_output_path)
            ws = wb.active
            current_row = 3

        def finalize_output_file(first_file, last_file):
            global combined_output_path
            def code_from_filename(filename, fallback):
                base = os.path.splitext(filename)[0] if filename else ""
                return base[:2].upper() if len(base) >= 2 else fallback

            def get_output_filename(counter, first_file, last_file):
                first_code = code_from_filename(first_file, "XX")
                last_code = code_from_filename(last_file, "YY")
                return f"{output_base_filename}_{counter}({first_code}-{last_code}).xlsx"

            wb.save(combined_output_path)
            final_name = get_output_filename(file_counter, first_file, last_file)
            final_path = os.path.join(output_dir, final_name)
            os.rename(combined_output_path, final_path)
            return final_path

        def clean_description(raw_description):
            raw_description = str(raw_description).strip()
            clean_desc = re.split(r'\s{5,}', raw_description)[0].strip()
            phrases = [p.strip() for p in clean_desc.split(';') if p.strip()]
            final_desc = ""
            for phrase in phrases:
                test_desc = phrase if not final_desc else f"{final_desc}; {phrase}"
                if len(test_desc) <= 40:
                    final_desc = test_desc
                else:
                    break
            if not final_desc and clean_desc:
                final_desc = clean_desc[:40]
            return final_desc

        # Start processing
        batch_first_file = uploaded_files[0].name
        batch_last_file = uploaded_files[0].name
        create_new_output_file()

        for uploaded in uploaded_files:
            file = uploaded.name
            st.write(f"🔍 Processing: {file}")
            try:
                if file.lower().endswith('.csv'):
                    df = pd.read_csv(uploaded, header=None, low_memory=False, keep_default_na=False, na_values=[])
                else:
                    df = pd.read_excel(uploaded, header=None, keep_default_na=False, na_values=[])
            except Exception as e:
                report.append(f"❌ Failed to read {file}: {e}")
                continue

            try:
                col_a = df.iloc[:, 0].astype(str)
                hashtag_row = 0 if col_a.iat[0] == "#" else col_a[col_a == "#"].index[0] if "#" in col_a.values else None
                if hashtag_row is None:
                    report.append(f"❌ '#' not found in column A of {file}")
                    continue

                if 10 >= len(df.columns):
                    report.append(f"❌ Column K does not exist in {file}")
                    continue

                col_k = df.iloc[:, 10].astype(str)
                dash_rows = col_k[col_k == "-"]
                if dash_rows.empty:
                    report.append(f"❌ '-' not found in column K of {file}")
                    continue

                start_row = hashtag_row + 1
                end_row = dash_rows.index[0]
                if start_row >= end_row:
                    report.append(f"⚠️ No data between '#' and '-' in {file}")
                    continue

                full_data_range = df.iloc[start_row:end_row, :]
                num_cols = full_data_range.shape[1]
                full_data_range.columns = column_names[:num_cols] + [f"Extra_{i}" for i in range(num_cols - len(column_names))]

                sin_number = sin_dict.get(file, "")

                for _, row in full_data_range.iterrows():
                    batch_last_file = file
                    if current_row > max_rows_per_file + 2:
                        finalize_output_file(batch_first_file, batch_last_file)
                        batch_first_file = file
                        file_counter += 1
                        create_new_output_file()

                    def write_cell(col, value):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = default_font
                        return cell

                    write_cell(1, "B")
                    write_cell(2, row.get("Manufacturer", ""))
                    part_number = row.get("Part Number", "")
                    write_cell(3, part_number)
                    write_cell(4, part_number)
                    write_cell(5, sin_number)
                    raw_description = row.get("Description", "")
                    cleaned_desc = clean_description(raw_description)
                    write_cell(6, cleaned_desc)
                    write_cell(7, raw_description)
                    write_cell(9, "EA")
                    write_cell(12, f"=P{current_row}*1.4")
                    write_cell(15, f"=P{current_row}*0.9925")
                    write_cell(16, row.get("Max List Price", ""))
                    write_cell(17, "MX")
                    write_cell(18, "15")
                    write_cell(19, "AE")
                    write_cell(20, "D")
                    write_cell(21, "O")
                    write_cell(22, "O")
                    write_cell(23, "O")
                    write_cell(30, "AlegnaLogo.jpg")
                    write_cell(32, row.get("Total Sales", ""))
                    current_row += 1
            except Exception as e:
                report.append(f"❌ Unexpected error in {file}: {e}")

        final_output = finalize_output_file(batch_first_file, batch_last_file)

        # Save processing report
        report_path = os.path.join(output_dir, "Processing_Report.txt")
        with open(report_path, "w") as f:
            for line in report:
                f.write(line + "\n")

        # Bundle results in ZIP for download
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for file in os.listdir(output_dir):
                zipf.write(os.path.join(output_dir, file), file)
        zip_buffer.seek(0)

        st.success("✅ Processing complete!")
        st.download_button("📥 Download Results (ZIP)", zip_buffer, file_name="Processed_Results.zip", mime="application/zip")
